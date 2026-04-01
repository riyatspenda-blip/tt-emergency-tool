import streamlit as st
import pandas as pd
from datetime import datetime
import io

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.title("TT Emergency Automation Tool")

uploaded_file = st.file_uploader(
    "Upload Excel Trouble Ticket",
    type=["xlsx"]
)

if uploaded_file:

    df = pd.read_excel(uploaded_file, sheet_name="Trouble Ticket Update")

    df_filtered = df[
        (df["CaseGroupName"] == "MMP/Intersite/Backhaul") &
        (df["ResolvedTimeOperator"].isna())
    ].copy()

    # =========================
    # WAKTU SEKARANG (FIX NO ERROR)
    # =========================
    now = datetime.now()  # ❗ TANPA timezone biar tidak error

    # =========================
    # FIX OPEN DATE (ANTI ERROR)
    # =========================
    df_filtered["OpenDate"] = pd.to_datetime(
        df_filtered["OpenDate"], errors="coerce"
    )

    # hilangkan timezone jika ada
    try:
        df_filtered["OpenDate"] = df_filtered["OpenDate"].dt.tz_localize(None)
    except:
        pass

    df_filtered = df_filtered.dropna(subset=["OpenDate"])

    # =========================
    # HITUNG DURASI
    # =========================
    delta = now - df_filtered["OpenDate"]

    df_filtered["durasi menit"] = (
        delta.dt.total_seconds() / 60
    ).fillna(0).astype(int)

    df_filtered["Duration"] = pd.to_timedelta(
        delta.dt.total_seconds(), unit="s"
    )

    # =========================
    # UPDATE TERAKHIR
    # =========================
    df_filtered["Update"] = df_filtered.get("LatestCIR", "")

    # =========================
    # CEK SLA
    # =========================
    def cek_sla(row):

        menit = row["durasi menit"]
        case = str(row["CaseName"]).lower()

        if case == "emergency" and menit <= 240:
            return "IN SLA"
        elif case == "major" and menit <= 1440:
            return "IN SLA"
        elif case == "minor" and menit <= 7200:
            return "IN SLA"
        else:
            return "OUT SLA"

    df_filtered["Ach. SLA Internal"] = df_filtered.apply(cek_sla, axis=1)

    # =========================
    # REMARK DURASI
    # =========================
    def remark(menit):

        if menit < 240:
            return "<4 jam"
        elif menit <= 480:
            return "4-8 jam"
        else:
            return ">8 jam"

    df_filtered["Remark durasi"] = df_filtered["durasi menit"].apply(remark)

    # =========================
    # MAPPING ROM
    # =========================
    rom_map = {
        "SULAWESI":"Abdul Karim",
        "SUMBAGSEL":"Eki Oktavian",
        "SUMBAGUT":"Charles Victor Steven Taneo",
        "KALIMANTAN":"Andri Potabuga",
        "JATIM":"Darwin",
        "BALINUSRA":"Anandayu Ega Hardianto",
        "JATENG":"Tri Pambudi",
        "JABAR":"Nonot Arief Herdianto",
        "SUMBAGTENG":"HARTONO",
        "JABODETABEK (OUTER)":"HENRO",
        "JABODETABEK (INNER)":"Chandra Novyan Nurfahmi",
        "LAMPUNG":"Fanel Situmorang Victor"
    }

    df_filtered["ROM"] = df_filtered["RegionName"].map(rom_map)

    # =========================
    # MFO OTOMATIS
    # =========================
    df_filtered["MFO"] = 0

    # =========================
    # KOLOM OUTPUT
    # =========================
    kolom_output = [
        "LogNo","CustomerTicketNo","SiteID","SiteName","ResidenceName",
        "CaseName","CaseDescription","OpenDate","SeverityName",
        "OperatorGroup","RegionName","VendorName","SPVOME","MFO",
        "Duration","durasi menit","Ach. SLA Internal","Update",
        "Remark durasi","ROM"
    ]

    for col in kolom_output:
        if col not in df_filtered.columns:
            df_filtered[col] = ""

    df_output = df_filtered[kolom_output]

    st.success("Data berhasil diproses")
    st.dataframe(df_output)

    # =========================
    # EXPORT EXCEL
    # =========================
    output = io.BytesIO()
    df_output.to_excel(output, index=False)
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active

    blue_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    red_header = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    header_font = Font(color="FFFFFF", bold=True)

    for col_num, cell in enumerate(ws[1], start=1):
        if col_num >= 15:
            cell.fill = red_header
        else:
            cell.fill = blue_header

        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # AUTO WIDTH
    for column in ws.columns:

        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        width = max_length + 3

        if column_letter in ["G","R"]:
            width = 40

        if width > 40:
            width = 40

        ws.column_dimensions[column_letter].width = width

    ws.auto_filter.ref = ws.dimensions

    thin = Side(style="thin")
    border = Border(left=thin,right=thin,top=thin,bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    # FORMAT DURATION
    for row in range(2, ws.max_row + 1):
        ws[f"O{row}"].number_format = "[h]:mm:ss"

    styled_output = io.BytesIO()
    wb.save(styled_output)

    # =========================
    # NAMA FILE DOWNLOAD (FIX)
    # =========================
    bulan = [
        "Januari","Februari","Maret","April","Mei","Juni",
        "Juli","Agustus","September","Oktober","November","Desember"
    ]

    now_download = datetime.now()

    tanggal = f"{now_download.day:02d} {bulan[now_download.month-1]} {now_download.year} {now_download.hour:02d}.{now_download.minute:02d}"

    filename = f"Trouble Ticket Emergency {tanggal}.xlsx"

    st.download_button(
        "Download Excel Report",
        styled_output.getvalue(),
        file_name=filename
    )
